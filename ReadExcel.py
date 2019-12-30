import openpyxl

vk = openpyxl.load_workbook("/home/adrianjimenez/Documents/robotframework/hoja1.xlsx")
print (vk.sheetnames)
print (vk.active.title)
workingSheet = vk['Hoja 5']
print (workingSheet.title)
print (workingSheet["A1"].value)

c1 = workingSheet.cell(2,1)
print (c1.value)

rows = workingSheet.max_row
columns = workingSheet.max_column

print ("number of rows is " + str(rows))
print ("number of columns is " + str(columns))
#print all values in the excel sheet
for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        c= workingSheet.cell(i,j)
        print (c.value)

#another approach
for r in workingSheet['A1':'D5']:
    for c in r:
        print (c.value)